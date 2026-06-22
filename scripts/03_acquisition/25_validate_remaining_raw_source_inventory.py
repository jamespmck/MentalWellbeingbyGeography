#!/usr/bin/env python3
"""
25_validate_remaining_raw_source_inventory.py

Validate and inventory remaining raw-source acquisition outputs for the
MentalWellbeingByGeography project.

This script does not join or process source data into analytical features.
It inspects the raw files staged by script 24 and creates an acquisition
validation register showing what is actually available, what each file appears
to contain, and what should happen next.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import logging
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

VERSION = "v14"
EXPECTED_PROJECT_ROOT = Path(r"D:\Good Measure\MentalWellbeingbyGeography")

GEOGRAPHY_PATTERNS = {
    "SA2": [r"\bsa2\b", r"statistical area level 2", r"sa2_?code"],
    "SA3": [r"\bsa3\b", r"statistical area level 3", r"sa3_?code"],
    "SA4": [r"\bsa4\b", r"statistical area level 4", r"sa4_?code"],
    "LGA": [r"\blga\b", r"local government area", r"lga_?code", r"council"],
    "PHN": [r"\bphn\b", r"primary health network", r"phn_?code"],
    "PHA": [r"\bpha\b", r"population health area", r"pha_?code"],
    "STATE_TERRITORY": [r"state", r"territory", r"jurisdiction", r"australia"],
    "HHS_LHD": [r"hospital and health service", r"local health district", r"\bhhs\b", r"\blhd\b"],
    "POSTCODE": [r"postcode", r"postal area", r"\bpoa\b"],
}

SOURCE_FAMILY_DEFAULT_SCOPE = {
    "aedc_child_development": "raw_inventory_then_geography_validation",
    "abs_homelessness_census": "raw_download_then_schema_geography_validation",
    "abs_homelessness_census_tablebuilder": "manual_access_or_tablebuilder_context",
    "aihw_specialist_homelessness_services": "raw_inventory_then_geography_validation",
    "aihw_mental_health_regional_activity": "raw_inventory_then_native_geography_table",
    "aihw_mental_health_data_tables": "raw_inventory_then_relevance_filter",
    "aihw_mbs_primary_care_geography": "raw_inventory_then_native_geography_table",
    "ndis_service_area_candidate": "raw_inventory_then_service_area_key_validation",
    "state_health_geography_inventory": "raw_inventory_only_state_specific",
}


def setup_logging(project_root: Path, debug: bool) -> Tuple[logging.Logger, Path]:
    logs_dir = project_root / "outputs" / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = logs_dir / f"25_validate_remaining_raw_source_inventory_{ts}.log"

    logger = logging.getLogger("remaining_raw_validation")
    logger.handlers.clear()
    logger.setLevel(logging.DEBUG if debug else logging.INFO)

    formatter = logging.Formatter("[%(asctime)s] [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
    stream = logging.StreamHandler(sys.stdout)
    stream.setFormatter(formatter)
    stream.setLevel(logging.DEBUG if debug else logging.INFO)
    logger.addHandler(stream)

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)

    return logger, log_path


def write_csv(path: Path, rows: List[Dict[str, Any]], logger: logging.Logger) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        # Write a one-row placeholder so downstream pandas reads do not fail.
        rows = [{"status": "empty", "notes": "No rows produced by this audit."}]
    logger.info("Writing CSV: %s", path)
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def safe_read_csv(path: Path, logger: logging.Logger) -> pd.DataFrame:
    if not path.exists():
        logger.warning("Missing CSV: %s", path)
        return pd.DataFrame()
    if path.stat().st_size == 0:
        logger.warning("Empty CSV: %s", path)
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except Exception as exc:
        logger.warning("Could not read CSV %s: %s", path, exc)
        return pd.DataFrame()


def normalise_path(value: Any) -> Optional[Path]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return None
    return Path(text)


def sha256_file(path: Path, max_bytes: Optional[int] = None) -> str:
    h = hashlib.sha256()
    read_bytes = 0
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            if max_bytes is not None and read_bytes + len(chunk) > max_bytes:
                chunk = chunk[: max_bytes - read_bytes]
            h.update(chunk)
            read_bytes += len(chunk)
            if max_bytes is not None and read_bytes >= max_bytes:
                break
    return h.hexdigest()


def extension_or_signature(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix:
        return suffix
    try:
        with path.open("rb") as f:
            head = f.read(16)
        if head.startswith(b"PK"):
            return ".zip_or_xlsx"
        if head.startswith(b"%PDF"):
            return ".pdf"
        if head.lstrip().startswith(b"<"):
            return ".html"
    except Exception:
        pass
    return "unknown"


def detect_geographies_from_text(text: str) -> str:
    haystack = text.lower()
    found: List[str] = []
    for geo, patterns in GEOGRAPHY_PATTERNS.items():
        if any(re.search(pattern, haystack, flags=re.IGNORECASE) for pattern in patterns):
            found.append(geo)
    return ";".join(found) if found else "unknown"


def safe_text_sample(value: Any, limit: int = 350) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text[:limit]


def inspect_text_file(path: Path) -> Dict[str, Any]:
    try:
        raw = path.read_bytes()[:250_000]
        text = raw.decode("utf-8", errors="ignore")
    except Exception as exc:
        return {"read_status": f"failed: {type(exc).__name__}: {exc}"}
    links = re.findall(r"https?://[^\s\"'<>]+", text, flags=re.IGNORECASE)
    return {
        "read_status": "ok",
        "text_sample": safe_text_sample(text, 500),
        "discovered_url_count_in_file": len(set(links)),
        "detected_geography_terms": detect_geographies_from_text(text),
    }


def inspect_zip(path: Path) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    members: List[Dict[str, Any]] = []
    try:
        with zipfile.ZipFile(path, "r") as zf:
            infos = zf.infolist()
            for info in infos:
                name = info.filename
                suffix = Path(name).suffix.lower()
                members.append(
                    {
                        "archive_path": str(path),
                        "member_name": name,
                        "member_extension": suffix,
                        "member_file_size": info.file_size,
                        "member_compress_size": info.compress_size,
                        "member_is_candidate_data": int(suffix in {".csv", ".xlsx", ".xls", ".json", ".txt"}),
                        "detected_geography_terms": detect_geographies_from_text(name),
                    }
                )
        summary = {
            "read_status": "ok",
            "zip_member_count": len(members),
            "zip_candidate_data_member_count": sum(r["member_is_candidate_data"] for r in members),
            "zip_member_extensions": ";".join(sorted({r["member_extension"] for r in members if r["member_extension"]})),
            "detected_geography_terms": detect_geographies_from_text(" ".join(r["member_name"] for r in members)),
        }
    except Exception as exc:
        summary = {"read_status": f"failed: {type(exc).__name__}: {exc}"}
    return summary, members


def sheet_header_sample(path: Path, sheet_name: str, max_rows: int = 25, max_cols: int = 60) -> Tuple[str, str, str, int, int]:
    """Return read_status, text_sample, detected_geographies, rows_read, cols_read."""
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=max_rows, engine="openpyxl")
        if df.shape[1] > max_cols:
            df = df.iloc[:, :max_cols]
        values = []
        for val in df.to_numpy().ravel().tolist():
            if pd.notna(val):
                values.append(str(val))
        text = " | ".join(values[:250])
        return "ok", safe_text_sample(text, 1000), detect_geographies_from_text(text), int(df.shape[0]), int(df.shape[1])
    except Exception as exc:
        return f"failed: {type(exc).__name__}: {exc}", "", "unknown", 0, 0


def inspect_workbook(path: Path, max_sheets: int, logger: logging.Logger) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    sheet_rows: List[Dict[str, Any]] = []
    try:
        if path.suffix.lower() == ".xls":
            # Modern pandas requires xlrd for xls. Try ExcelFile and record failure if xlrd absent.
            xls = pd.ExcelFile(path)
            sheet_names = list(xls.sheet_names)
            engine = getattr(xls, "engine", "unknown")
        else:
            if load_workbook is not None:
                wb = load_workbook(path, read_only=True, data_only=True)
                sheet_names = list(wb.sheetnames)
                engine = "openpyxl"
                # Use openpyxl dimensions without loading whole sheets.
                dimensions: Dict[str, Tuple[Optional[int], Optional[int]]] = {}
                for s in sheet_names[:max_sheets]:
                    ws = wb[s]
                    dimensions[s] = (ws.max_row, ws.max_column)
                wb.close()
            else:
                xls = pd.ExcelFile(path)
                sheet_names = list(xls.sheet_names)
                engine = getattr(xls, "engine", "unknown")
                dimensions = {}

        selected = sheet_names[:max_sheets]
        for s in selected:
            read_status, sample, geos, rows_read, cols_read = sheet_header_sample(path, s)
            max_row = None
            max_col = None
            if "dimensions" in locals() and s in dimensions:
                max_row, max_col = dimensions[s]
            sheet_rows.append(
                {
                    "workbook_path": str(path),
                    "workbook_name": path.name,
                    "sheet_name": s,
                    "sheet_index": sheet_names.index(s),
                    "sheet_rows_reported": max_row,
                    "sheet_columns_reported": max_col,
                    "sample_rows_read": rows_read,
                    "sample_columns_read": cols_read,
                    "sample_read_status": read_status,
                    "detected_geography_terms": geos,
                    "header_text_sample": sample,
                }
            )
        detected = detect_geographies_from_text(" ".join([path.name] + sheet_names))
        summary = {
            "read_status": "ok",
            "workbook_engine": engine,
            "sheet_count": len(sheet_names),
            "sheets_sampled": len(selected),
            "sheet_names_sample": " | ".join(sheet_names[:25]),
            "detected_geography_terms": detected,
        }
    except Exception as exc:
        logger.debug("Workbook inspect failed for %s: %s", path, exc)
        summary = {"read_status": f"failed: {type(exc).__name__}: {exc}"}
    return summary, sheet_rows


def derive_next_action(source_family: str, file_type: str, detected_geos: str, read_status: str, raw_path: str) -> str:
    if read_status.startswith("failed"):
        return "manual_review_file_unreadable_or_requires_special_parser"
    if source_family == "state_health_geography_inventory":
        return "hold_state_specific_boundary_context_only"
    if source_family == "abs_homelessness_census":
        if "SA2" in detected_geos or "LGA" in detected_geos or "STATE_TERRITORY" in detected_geos:
            return "process_abs_homelessness_schema_and_native_geography"
        return "inspect_abs_homelessness_workbook_tables"
    if source_family == "aedc_child_development":
        return "inspect_aedc_geography_and_release_scope"
    if source_family == "aihw_mental_health_regional_activity":
        return "unzip_and_inventory_aihw_mental_health_tables"
    if source_family == "aihw_specialist_homelessness_services":
        return "inspect_shs_public_tables_for_geography_or_hold_report_context"
    if source_family == "aihw_mbs_primary_care_geography":
        return "inspect_mbs_local_area_report_and_downloadable_tables"
    if source_family == "ndis_service_area_candidate":
        return "hold_until_true_ndis_service_area_key_identified"
    if file_type in {".zip", ".zip_or_xlsx"}:
        return "unzip_and_inventory_candidate_tables"
    return SOURCE_FAMILY_DEFAULT_SCOPE.get(source_family, "manual_review")


def collect_raw_file_records(register: pd.DataFrame, link_audit: pd.DataFrame) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def add_rows(df: pd.DataFrame, origin: str) -> None:
        if df.empty or "raw_file_path" not in df.columns:
            return
        for idx, row in df.iterrows():
            raw_path = normalise_path(row.get("raw_file_path"))
            if raw_path is None:
                continue
            rec = {"record_origin": origin, "record_index": idx}
            for col in [
                "source_family",
                "publisher",
                "source_name",
                "url",
                "page_url",
                "discovered_url",
                "download_status",
                "http_status",
                "content_type",
                "recommended_scope",
                "notes",
                "link_text",
                "extension",
            ]:
                if col in df.columns:
                    rec[col] = row.get(col)
            rec["raw_file_path"] = str(raw_path)
            rows.append(rec)

    add_rows(register, "register")
    add_rows(link_audit, "candidate_link_audit")

    # Deduplicate by path, while preserving combined source-family information.
    dedup: Dict[str, Dict[str, Any]] = {}
    for rec in rows:
        key = rec["raw_file_path"]
        if key not in dedup:
            dedup[key] = rec
        else:
            existing = dedup[key]
            existing["record_origin"] = f"{existing.get('record_origin')};{rec.get('record_origin')}"
            for col in ["source_family", "source_name", "download_status"]:
                if not existing.get(col) and rec.get(col):
                    existing[col] = rec[col]
    return list(dedup.values())


def validate_files(project_root: Path, logger: logging.Logger, args: argparse.Namespace) -> None:
    audits_dir = project_root / "outputs" / "audits"
    docs_register_dir = project_root / "docs" / "source_registers"
    docs_method_dir = project_root / "docs" / "methodology"
    for p in [audits_dir, docs_register_dir, docs_method_dir]:
        p.mkdir(parents=True, exist_ok=True)

    register_path = audits_dir / "remaining_raw_source_acquisition_register_v13.csv"
    links_path = audits_dir / "remaining_raw_source_candidate_link_audit_v13.csv"

    register = safe_read_csv(register_path, logger)
    links = safe_read_csv(links_path, logger)

    file_records = collect_raw_file_records(register, links)
    logger.info("Raw file records to validate: %s", len(file_records))

    file_inventory: List[Dict[str, Any]] = []
    workbook_sheets: List[Dict[str, Any]] = []
    zip_members: List[Dict[str, Any]] = []

    for i, rec in enumerate(file_records, start=1):
        path = Path(rec["raw_file_path"])
        source_family = str(rec.get("source_family", "unknown"))
        logger.info("[%s/%s] Inspecting %s", i, len(file_records), path.name)

        base: Dict[str, Any] = {
            "run_timestamp": datetime.now().isoformat(timespec="seconds"),
            "source_family": source_family,
            "publisher": rec.get("publisher"),
            "source_name": rec.get("source_name"),
            "record_origin": rec.get("record_origin"),
            "download_status_from_source_register": rec.get("download_status"),
            "source_url": rec.get("url") or rec.get("page_url") or rec.get("discovered_url"),
            "raw_file_path": str(path),
            "file_exists": int(path.exists()),
            "file_type": extension_or_signature(path) if path.exists() else "missing",
            "file_size_bytes": path.stat().st_size if path.exists() else 0,
            "sha256": sha256_file(path) if path.exists() else None,
        }

        detail: Dict[str, Any] = {}
        if not path.exists():
            detail = {"read_status": "missing"}
        else:
            ftype = base["file_type"]
            if ftype in {".xlsx", ".xls"}:
                summary, sheets = inspect_workbook(path, args.max_workbook_sheets, logger)
                detail.update(summary)
                workbook_sheets.extend(sheets)
            elif ftype == ".zip":
                summary, members = inspect_zip(path)
                detail.update(summary)
                zip_members.extend(members)
            elif ftype in {".html", ".download", ".txt"} or str(path).lower().endswith(".download"):
                detail.update(inspect_text_file(path))
            elif ftype == ".pdf":
                detail.update({"read_status": "binary_pdf_not_parsed", "detected_geography_terms": detect_geographies_from_text(path.name)})
            else:
                # Try text as fallback.
                detail.update(inspect_text_file(path))

        detected = detail.get("detected_geography_terms") or detect_geographies_from_text(path.name)
        read_status = str(detail.get("read_status", "unknown"))
        base.update(detail)
        base["native_geography_detected"] = detected
        base["recommended_next_action"] = derive_next_action(source_family, str(base.get("file_type")), str(detected), read_status, str(path))
        file_inventory.append(base)

    # Summary by source family.
    summary_rows: List[Dict[str, Any]] = []
    if file_inventory:
        inv_df = pd.DataFrame(file_inventory)
        grouped = inv_df.groupby("source_family", dropna=False)
        for sf, g in grouped:
            existing = int(g["file_exists"].sum()) if "file_exists" in g else 0
            summary_rows.append(
                {
                    "source_family": sf,
                    "raw_files_seen": int(len(g)),
                    "raw_files_existing": existing,
                    "file_types": ";".join(sorted(set(str(x) for x in g.get("file_type", pd.Series(dtype=str)).dropna()))),
                    "detected_geographies": ";".join(sorted(set(";".join(g.get("native_geography_detected", pd.Series(dtype=str)).fillna("unknown")).split(";")))) if "native_geography_detected" in g else "unknown",
                    "recommended_next_actions": ";".join(sorted(set(str(x) for x in g.get("recommended_next_action", pd.Series(dtype=str)).dropna()))),
                    "validation_status": "pass_has_raw_files" if existing > 0 else "manual_review_no_raw_files",
                }
            )

    # Add acquisition-register failed rows for visibility.
    failed_rows: List[Dict[str, Any]] = []
    if not register.empty:
        mask = register.astype(str).apply(lambda r: r.str.contains("fail|error|forbidden|exception", case=False, na=False).any(), axis=1)
        for _, r in register[mask].iterrows():
            failed_rows.append({k: r.get(k) for k in register.columns})

    write_csv(audits_dir / "remaining_raw_source_file_inventory_v14.csv", file_inventory, logger)
    write_csv(audits_dir / "remaining_raw_workbook_sheet_inventory_v14.csv", workbook_sheets, logger)
    write_csv(audits_dir / "remaining_raw_zip_member_inventory_v14.csv", zip_members, logger)
    write_csv(audits_dir / "remaining_raw_source_validation_summary_v14.csv", summary_rows, logger)
    write_csv(audits_dir / "remaining_raw_source_failed_register_rows_v14.csv", failed_rows, logger)
    write_csv(docs_register_dir / "remaining_raw_source_validation_register_v14.csv", file_inventory, logger)

    note = f"""# Remaining raw source validation note {VERSION}

Generated: {datetime.now().isoformat(timespec='seconds')}

This validation pass inspects raw files staged by `24_acquire_remaining_raw_source_register.py`.
It does not join sources and does not create modelling features.

## Outputs

- `outputs/audits/remaining_raw_source_file_inventory_v14.csv`
- `outputs/audits/remaining_raw_workbook_sheet_inventory_v14.csv`
- `outputs/audits/remaining_raw_zip_member_inventory_v14.csv`
- `outputs/audits/remaining_raw_source_validation_summary_v14.csv`
- `outputs/audits/remaining_raw_source_failed_register_rows_v14.csv`
- `docs/source_registers/remaining_raw_source_validation_register_v14.csv`

## Interpretation

Use the validation summary to decide which raw sources are ready for native-geography processing.
State-specific health geography sources should remain context only unless a validated crosswalk is available.
Higher-level sources such as LGA and PHN should be kept separate from the SA2 master and connected later through the scoped foreign-key model.
"""
    note_path = docs_method_dir / "remaining_raw_source_validation_note_v14.md"
    logger.info("Writing methodology note: %s", note_path)
    note_path.write_text(note, encoding="utf-8")

    logger.info("Remaining raw source validation complete.")
    logger.info("Summary rows: %s", len(summary_rows))
    if failed_rows:
        logger.info("Failed acquisition rows carried forward: %s", len(failed_rows))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate remaining raw-source acquisition outputs.")
    parser.add_argument("--project-root", default=str(EXPECTED_PROJECT_ROOT), help="Project root path.")
    parser.add_argument("--debug", action="store_true", help="Enable verbose logging.")
    parser.add_argument("--max-workbook-sheets", type=int, default=80, help="Maximum workbook sheets to sample per workbook.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    project_root = Path(args.project_root)
    logger, log_path = setup_logging(project_root, args.debug)
    logger.info("Remaining raw source validation %s", VERSION)
    logger.info("Project root: %s", project_root)
    logger.info("Log path: %s", log_path)
    validate_files(project_root, logger, args)


if __name__ == "__main__":
    main()
