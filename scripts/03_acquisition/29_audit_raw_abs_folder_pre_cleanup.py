#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Audit data/raw/abs before cleanup.

Read-only script. It inventories files, detects duplicates, inspects workbook sheets,
zip members and HTML/source-page text, and writes cleanup-candidate CSVs.

Designed for:
  D:\Good Measure\MentalWellbeingbyGeography

Usage:
  python scripts\03_acquisition\29_audit_raw_abs_folder_pre_cleanup.py --debug
  python scripts\03_acquisition\29_audit_raw_abs_folder_pre_cleanup.py --target-dir "D:\Good Measure\MentalWellbeingbyGeography\data\raw\abs" --debug
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    import pandas as pd
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"pandas is required for this audit script: {exc}")

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    import xlrd  # type: ignore
except Exception:
    xlrd = None

TEXT_EXTENSIONS = {".txt", ".csv", ".tsv", ".json", ".html", ".htm", ".xml", ".md"}
WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
ARCHIVE_EXTENSIONS = {".zip"}
BINARY_EXTENSIONS = {".pdf", ".parquet", ".feather", ".png", ".jpg", ".jpeg", ".gif", ".webp"}

GEOGRAPHY_PATTERNS = {
    "SA1": [r"\bSA1\b", r"sa1[_\s-]*code", r"statistical area level 1"],
    "SA2": [r"\bSA2\b", r"sa2[_\s-]*code", r"statistical area level 2"],
    "SA3": [r"\bSA3\b", r"sa3[_\s-]*code", r"statistical area level 3"],
    "SA4": [r"\bSA4\b", r"sa4[_\s-]*code", r"statistical area level 4"],
    "LGA": [r"\bLGA\b", r"local government area"],
    "PHN": [r"\bPHN\b", r"primary health network"],
    "STATE_TERRITORY": [r"state", r"territor", r"jurisdiction", r"\bNSW\b", r"\bVIC\b", r"\bQLD\b", r"\bSA\b", r"\bWA\b", r"\bTAS\b", r"\bNT\b", r"\bACT\b"],
    "REMOTENESS": [r"remoteness", r"major cit", r"inner regional", r"outer regional", r"remote", r"very remote"],
    "CENSUS": [r"census", r"quickstats", r"tablebuilder", r"datapack"],
    "HOMELESSNESS": [r"homeless", r"marginal", r"boarding house", r"rough sleeping", r"improvised dwelling"],
}

SOURCE_HINT_PATTERNS = {
    "abs_homelessness_census": [r"homeless", r"20490", r"estimating_homelessness"],
    "abs_tablebuilder_homelessness": [r"tablebuilder", r"2049055002"],
    "abs_census_quickstats": [r"quickstats", r"2021census", r"gcp", r"datapack"],
    "abs_geography_bridge": [r"correspondence", r"allocation", r"asgs", r"sa2_2021", r"lga", r"phn"],
    "abs_seifa": [r"seifa", r"irsd", r"irsad", r"socio-economic indexes"],
    "abs_remoteness": [r"remoteness", r"ra_2021"],
}


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def log(msg: str, debug: bool = True) -> None:
    if debug:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{ts}] {msg}")


def find_project_root(start: Path) -> Path:
    candidates = [start, *start.parents]
    for p in candidates:
        if (p / "data").exists() and (p / "scripts").exists():
            return p
    # Common fallback when script is run from project root.
    return Path.cwd()


def sha256_file(path: Path, block_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            b = f.read(block_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def safe_rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)


def read_text_sample(path: Path, max_bytes: int = 100_000) -> str:
    try:
        raw = path.read_bytes()[:max_bytes]
    except Exception:
        return ""
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
        try:
            txt = raw.decode(enc, errors="replace")
            txt = html.unescape(txt)
            txt = re.sub(r"\s+", " ", txt).strip()
            return txt[:5000]
        except Exception:
            continue
    return ""


def detect_geographies(text: str) -> str:
    if not text:
        return "unknown"
    found: list[str] = []
    for label, pats in GEOGRAPHY_PATTERNS.items():
        if any(re.search(p, text, flags=re.IGNORECASE) for p in pats):
            found.append(label)
    return ";".join(found) if found else "unknown"


def detect_source_hint(path: Path, text: str) -> str:
    hay = f"{path.name} {path.parent.name} {text[:1000]}"
    found: list[str] = []
    for label, pats in SOURCE_HINT_PATTERNS.items():
        if any(re.search(p, hay, flags=re.IGNORECASE) for p in pats):
            found.append(label)
    return ";".join(found) if found else "unknown"


def count_html_links(text: str) -> int:
    if not text:
        return 0
    return len(re.findall(r"href\s*=\s*['\"]", text, flags=re.IGNORECASE))


def inspect_xlsx(path: Path, max_sample_sheets: int = 80) -> tuple[str, int | None, str, list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    if load_workbook is None:
        return "openpyxl_not_available", None, "", rows
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        for ws in wb.worksheets[:max_sample_sheets]:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            # Collect first few non-empty cells to assist later cleanup decisions.
            sample_values: list[str] = []
            inspected = 0
            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 12), values_only=True):
                inspected += 1
                for val in row[: min(max_col, 12)]:
                    if val is None:
                        continue
                    s = str(val).strip()
                    if s:
                        sample_values.append(s[:120])
                    if len(sample_values) >= 20:
                        break
                if len(sample_values) >= 20:
                    break
            sample_text = " | ".join(sample_values)
            rows.append({
                "file_path": str(path),
                "file_name": path.name,
                "sheet_name": ws.title,
                "sheet_index": sheet_names.index(ws.title) + 1,
                "max_row": max_row,
                "max_column": max_col,
                "first_cells_sample": sample_text,
                "detected_geographies": detect_geographies(sample_text),
                "likely_blank_or_cover_sheet": int(max_row <= 5 or max_col <= 2),
            })
        wb.close()
        return "ok", len(sheet_names), " | ".join(sheet_names[:40]), rows
    except Exception as exc:
        return f"error:{type(exc).__name__}:{exc}", None, "", rows


def inspect_xls(path: Path, max_sample_sheets: int = 80) -> tuple[str, int | None, str, list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    if xlrd is None:
        return "xlrd_not_available", None, "", rows
    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
        sheet_names = book.sheet_names()
        for i, sh_name in enumerate(sheet_names[:max_sample_sheets], start=1):
            sh = book.sheet_by_name(sh_name)
            sample_values: list[str] = []
            for r in range(min(sh.nrows, 12)):
                for c in range(min(sh.ncols, 12)):
                    val = sh.cell_value(r, c)
                    if val is None:
                        continue
                    s = str(val).strip()
                    if s:
                        sample_values.append(s[:120])
                    if len(sample_values) >= 20:
                        break
                if len(sample_values) >= 20:
                    break
            sample_text = " | ".join(sample_values)
            rows.append({
                "file_path": str(path),
                "file_name": path.name,
                "sheet_name": sh_name,
                "sheet_index": i,
                "max_row": sh.nrows,
                "max_column": sh.ncols,
                "first_cells_sample": sample_text,
                "detected_geographies": detect_geographies(sample_text),
                "likely_blank_or_cover_sheet": int(sh.nrows <= 5 or sh.ncols <= 2),
            })
        book.release_resources()
        return "ok", len(sheet_names), " | ".join(sheet_names[:40]), rows
    except Exception as exc:
        return f"error:{type(exc).__name__}:{exc}", None, "", rows


def inspect_zip(path: Path) -> tuple[str, int | None, int | None, str, list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(path, "r") as zf:
            members = zf.infolist()
            candidates = 0
            extensions = Counter()
            for info in members:
                ext = Path(info.filename).suffix.lower()
                if ext:
                    extensions[ext] += 1
                is_candidate = ext in {".csv", ".xlsx", ".xls", ".txt", ".json", ".parquet"}
                candidates += int(is_candidate)
                rows.append({
                    "archive_path": str(path),
                    "archive_name": path.name,
                    "member_name": info.filename,
                    "member_extension": ext,
                    "member_file_size": info.file_size,
                    "member_compress_size": info.compress_size,
                    "member_is_dir": int(info.is_dir()),
                    "member_is_candidate_data": int(is_candidate),
                    "detected_geographies": detect_geographies(info.filename),
                })
            ext_summary = ";".join([f"{k}:{v}" for k, v in sorted(extensions.items())])
            return "ok", len(members), candidates, ext_summary, rows
    except Exception as exc:
        return f"error:{type(exc).__name__}:{exc}", None, None, "", rows


def candidate_cleanup_action(record: dict[str, Any]) -> tuple[str, str]:
    """Return (candidate_action, reason). Read-only recommendation only."""
    rel = record.get("relative_path", "")
    name = record.get("file_name", "")
    parent = record.get("parent_folder", "")
    ext = record.get("file_type", "")
    source_hint = record.get("source_hint", "")
    size = int(record.get("file_size_bytes", 0) or 0)
    duplicate_group_size = int(record.get("duplicate_group_size", 1) or 1)
    path_lower = rel.lower().replace("\\", "/")
    name_lower = name.lower()

    if duplicate_group_size > 1:
        return "review_duplicate_keep_one_active_copy", "Same sha256 appears in multiple ABS raw locations."
    if "browser" in path_lower or "capture" in path_lower or "network" in path_lower or "screenshot" in path_lower:
        return "archive_browser_capture_debris", "Browser/debug capture artefact should not remain in active raw source layer."
    if ext in {".tmp", ".crdownload", ".part"}:
        return "delete_or_archive_partial_download", "Partial/temporary download extension."
    if ext in BINARY_EXTENSIONS and ext != ".pdf":
        return "archive_non_data_binary", "Non-data binary file in ABS raw folder."
    if "duplicate" in name_lower or "copy" in name_lower:
        return "review_named_duplicate", "Filename suggests a duplicate/root copy."
    if ext in {".html", ".htm"}:
        return "keep_source_page_snapshot", "HTML source pages are provenance snapshots."
    if ext in WORKBOOK_EXTENSIONS:
        return "keep_pending_schema_inspection", "Workbook is a candidate source table pending native-geography processing."
    if ext in {".csv", ".tsv"}:
        return "keep_pending_schema_inspection", "Delimited file is a candidate source table pending native-geography processing."
    if ext == ".zip":
        return "keep_or_extract_after_member_review", "Archive should be kept until member inventory is processed."
    if size == 0:
        return "archive_or_delete_zero_byte_file", "Zero-byte file."
    if source_hint == "unknown":
        return "manual_review_unknown_abs_raw_file", "ABS source family could not be inferred from filename/path/text sample."
    return "keep", "No cleanup issue detected."


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    df = pd.DataFrame(rows)
    df.to_csv(path, index=False)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Audit data/raw/abs before cleanup. Read-only.")
    parser.add_argument("--project-root", default=None, help="Project root. Defaults to inferred/current directory.")
    parser.add_argument("--target-dir", default=None, help="Folder to audit. Defaults to data/raw/abs under project root.")
    parser.add_argument("--debug", action="store_true", help="Print progress.")
    parser.add_argument("--hash-large-files", action="store_true", help="Hash files larger than 1 GB. Default skips hashes for >1 GB.")
    args = parser.parse_args(argv)

    project_root = Path(args.project_root).resolve() if args.project_root else find_project_root(Path.cwd().resolve())
    target_dir = Path(args.target_dir).resolve() if args.target_dir else (project_root / "data" / "raw" / "abs")
    stamp = now_stamp()

    audit_dir = project_root / "outputs" / "audits"
    note_dir = project_root / "docs" / "methodology"
    audit_dir.mkdir(parents=True, exist_ok=True)
    note_dir.mkdir(parents=True, exist_ok=True)

    file_inventory_path = audit_dir / f"raw_abs_file_inventory_v18_{stamp}.csv"
    duplicate_path = audit_dir / f"raw_abs_duplicate_groups_v18_{stamp}.csv"
    workbook_path = audit_dir / f"raw_abs_workbook_sheet_inventory_v18_{stamp}.csv"
    zip_path = audit_dir / f"raw_abs_zip_member_inventory_v18_{stamp}.csv"
    folder_summary_path = audit_dir / f"raw_abs_folder_summary_v18_{stamp}.csv"
    cleanup_path = audit_dir / f"raw_abs_cleanup_candidate_audit_v18_{stamp}.csv"
    run_path = audit_dir / f"raw_abs_audit_run_summary_v18_{stamp}.csv"
    note_path = note_dir / f"raw_abs_folder_audit_note_v18_{stamp}.md"

    log(f"Project root: {project_root}", args.debug)
    log(f"Target dir:   {target_dir}", args.debug)

    run_rows: list[dict[str, Any]] = []
    if not target_dir.exists():
        # Give useful diagnostics without failing silently.
        sibling_candidates = sorted((project_root / "data" / "raw").glob("abs*")) if (project_root / "data" / "raw").exists() else []
        run_rows.append({
            "run_timestamp": stamp,
            "project_root": str(project_root),
            "target_dir": str(target_dir),
            "target_exists": 0,
            "file_count": 0,
            "folder_count": 0,
            "total_size_bytes": 0,
            "status": "target_dir_missing",
            "sibling_abs_candidates": ";".join(str(p) for p in sibling_candidates),
        })
        write_csv(run_path, run_rows)
        note_path.write_text("\n".join([
            "# Raw ABS folder audit note",
            "",
            f"Run timestamp: {stamp}",
            f"Project root: {project_root}",
            f"Target directory: {target_dir}",
            "",
            "Status: target directory missing.",
            "",
            "Sibling candidates found:",
            *[f"- {p}" for p in sibling_candidates],
            "",
        ]), encoding="utf-8")
        print("Target directory not found.")
        print(f"Expected: {target_dir}")
        if sibling_candidates:
            print("Possible ABS raw folders:")
            for p in sibling_candidates:
                print(f"  {p}")
        print(f"Run summary: {run_path}")
        print(f"Note: {note_path}")
        return 2

    files = [p for p in sorted(target_dir.rglob("*")) if p.is_file()]
    dirs = [p for p in sorted(target_dir.rglob("*")) if p.is_dir()]
    log(f"Files found: {len(files)}", args.debug)

    file_rows: list[dict[str, Any]] = []
    workbook_rows: list[dict[str, Any]] = []
    zip_rows: list[dict[str, Any]] = []

    for i, path in enumerate(files, start=1):
        rel = safe_rel(path, project_root)
        rel_target = safe_rel(path, target_dir)
        ext = path.suffix.lower()
        size = path.stat().st_size
        log(f"[{i}/{len(files)}] {rel_target}", args.debug)

        read_status = "not_read"
        text_sample = ""
        link_count = None
        sheet_count = None
        sheet_names_sample = ""
        zip_member_count = None
        zip_candidate_data_member_count = None
        zip_member_extensions = ""
        geography_text = ""

        # Hash, skipping very large files unless requested.
        if size > 1_000_000_000 and not args.hash_large_files:
            file_hash = "SKIPPED_GT_1GB_USE_HASH_LARGE_FILES"
        else:
            try:
                file_hash = sha256_file(path)
            except Exception as exc:
                file_hash = f"HASH_ERROR:{type(exc).__name__}:{exc}"

        if ext in TEXT_EXTENSIONS:
            text_sample = read_text_sample(path)
            read_status = "ok" if text_sample else "empty_or_unreadable_text"
            if ext in {".html", ".htm"}:
                link_count = count_html_links(text_sample)
            geography_text = text_sample
        elif ext in {".xlsx", ".xlsm"}:
            read_status, sheet_count, sheet_names_sample, rows = inspect_xlsx(path)
            workbook_rows.extend(rows)
            geography_text = f"{path.name} {sheet_names_sample} " + " ".join(r.get("first_cells_sample", "") for r in rows[:5])
        elif ext == ".xls":
            read_status, sheet_count, sheet_names_sample, rows = inspect_xls(path)
            workbook_rows.extend(rows)
            geography_text = f"{path.name} {sheet_names_sample} " + " ".join(r.get("first_cells_sample", "") for r in rows[:5])
        elif ext == ".zip":
            read_status, zip_member_count, zip_candidate_data_member_count, zip_member_extensions, rows = inspect_zip(path)
            zip_rows.extend(rows)
            geography_text = f"{path.name} " + " ".join(r.get("member_name", "") for r in rows[:50])
        elif ext in BINARY_EXTENSIONS:
            read_status = "binary_not_parsed"
            geography_text = path.name
        else:
            text_sample = read_text_sample(path, max_bytes=20_000)
            read_status = "unknown_type_text_sampled" if text_sample else "unknown_type_not_read"
            geography_text = f"{path.name} {text_sample}"

        detected_geographies = detect_geographies(geography_text)
        source_hint = detect_source_hint(path, geography_text)

        file_rows.append({
            "run_timestamp": stamp,
            "project_root": str(project_root),
            "target_dir": str(target_dir),
            "file_path": str(path),
            "relative_path": rel,
            "relative_path_within_target": rel_target,
            "file_name": path.name,
            "parent_folder": path.parent.name,
            "depth_within_target": len(Path(rel_target).parts) - 1,
            "file_type": ext if ext else "no_extension",
            "file_size_bytes": size,
            "file_size_mb": round(size / (1024 * 1024), 6),
            "modified_time": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
            "sha256": file_hash,
            "read_status": read_status,
            "html_link_count": link_count,
            "workbook_sheet_count": sheet_count,
            "workbook_sheet_names_sample": sheet_names_sample,
            "zip_member_count": zip_member_count,
            "zip_candidate_data_member_count": zip_candidate_data_member_count,
            "zip_member_extensions": zip_member_extensions,
            "detected_geographies": detected_geographies,
            "source_hint": source_hint,
            "text_sample": text_sample[:1000],
        })

    # Duplicate groups by hash.
    hash_counts = Counter(r["sha256"] for r in file_rows if r.get("sha256") and not str(r["sha256"]).startswith(("HASH_ERROR", "SKIPPED")))
    for r in file_rows:
        r["duplicate_group_size"] = hash_counts.get(r.get("sha256"), 0)
        action, reason = candidate_cleanup_action(r)
        r["cleanup_candidate_action"] = action
        r["cleanup_candidate_reason"] = reason

    duplicate_rows: list[dict[str, Any]] = []
    dup_group_id = 0
    by_hash: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in file_rows:
        h = str(r.get("sha256", ""))
        if h and not h.startswith(("HASH_ERROR", "SKIPPED")):
            by_hash[h].append(r)
    for h, group in by_hash.items():
        if len(group) <= 1:
            continue
        dup_group_id += 1
        # Suggested keep: shortest path and not inside archive/quarantine/debug.
        def keep_score(row: dict[str, Any]) -> tuple[int, int, str]:
            rel = str(row.get("relative_path", "")).lower()
            penalty = 0
            if "archive" in rel or "_archive" in rel or "quarantine" in rel or "debug" in rel or "browser" in rel:
                penalty += 100
            if "discovered_downloads" in rel:
                penalty += 10
            return (penalty, len(rel), rel)
        suggested = sorted(group, key=keep_score)[0]
        for row in sorted(group, key=lambda x: str(x.get("relative_path", ""))):
            duplicate_rows.append({
                "duplicate_group_id": dup_group_id,
                "sha256": h,
                "group_size": len(group),
                "file_path": row["file_path"],
                "relative_path": row["relative_path"],
                "file_name": row["file_name"],
                "file_size_bytes": row["file_size_bytes"],
                "suggested_keep": int(row["file_path"] == suggested["file_path"]),
                "suggested_action": "keep_active_copy" if row["file_path"] == suggested["file_path"] else "archive_duplicate_copy",
            })

    # Folder summary.
    folder_rows: list[dict[str, Any]] = []
    folder_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in file_rows:
        folder_groups[str(Path(r["relative_path_within_target"]).parent)].append(r)
    for folder, rows in sorted(folder_groups.items()):
        folder_rows.append({
            "folder_within_target": folder,
            "file_count": len(rows),
            "total_size_bytes": sum(int(r["file_size_bytes"]) for r in rows),
            "total_size_mb": round(sum(int(r["file_size_bytes"]) for r in rows) / (1024 * 1024), 6),
            "file_types": ";".join(sorted(set(str(r["file_type"]) for r in rows))),
            "source_hints": ";".join(sorted(set(str(r["source_hint"]) for r in rows))),
            "cleanup_actions": ";".join(sorted(set(str(r["cleanup_candidate_action"]) for r in rows))),
        })

    cleanup_rows = []
    for r in file_rows:
        if r["cleanup_candidate_action"] != "keep":
            cleanup_rows.append({
                "relative_path": r["relative_path"],
                "file_name": r["file_name"],
                "file_type": r["file_type"],
                "file_size_mb": r["file_size_mb"],
                "sha256": r["sha256"],
                "duplicate_group_size": r["duplicate_group_size"],
                "source_hint": r["source_hint"],
                "detected_geographies": r["detected_geographies"],
                "cleanup_candidate_action": r["cleanup_candidate_action"],
                "cleanup_candidate_reason": r["cleanup_candidate_reason"],
            })

    write_csv(file_inventory_path, file_rows)
    write_csv(duplicate_path, duplicate_rows)
    write_csv(workbook_path, workbook_rows)
    write_csv(zip_path, zip_rows)
    write_csv(folder_summary_path, folder_rows)
    write_csv(cleanup_path, cleanup_rows)

    run_rows.append({
        "run_timestamp": stamp,
        "project_root": str(project_root),
        "target_dir": str(target_dir),
        "target_exists": 1,
        "file_count": len(files),
        "folder_count": len(dirs),
        "total_size_bytes": sum(p.stat().st_size for p in files),
        "total_size_mb": round(sum(p.stat().st_size for p in files) / (1024 * 1024), 6),
        "duplicate_groups": dup_group_id,
        "duplicate_file_rows": len(duplicate_rows),
        "workbook_files": sum(1 for r in file_rows if r["file_type"] in WORKBOOK_EXTENSIONS),
        "zip_files": sum(1 for r in file_rows if r["file_type"] == ".zip"),
        "html_files": sum(1 for r in file_rows if r["file_type"] in {".html", ".htm"}),
        "cleanup_candidate_rows": len(cleanup_rows),
        "status": "pass_audit_written",
        "sibling_abs_candidates": "",
    })
    write_csv(run_path, run_rows)

    note_lines = [
        "# Raw ABS folder audit note",
        "",
        f"Run timestamp: {stamp}",
        f"Project root: {project_root}",
        f"Target directory: {target_dir}",
        "",
        "This was a read-only audit. No files were moved, renamed or deleted.",
        "",
        "## Outputs",
        f"- {file_inventory_path}",
        f"- {duplicate_path}",
        f"- {workbook_path}",
        f"- {zip_path}",
        f"- {folder_summary_path}",
        f"- {cleanup_path}",
        f"- {run_path}",
        "",
        "## Run summary",
        f"- Files: {len(files)}",
        f"- Folders: {len(dirs)}",
        f"- Total size MB: {run_rows[0]['total_size_mb']}",
        f"- Duplicate groups: {dup_group_id}",
        f"- Cleanup candidate rows: {len(cleanup_rows)}",
        "",
        "Review the cleanup candidate audit before any cleanup script is generated.",
        "",
    ]
    note_path.write_text("\n".join(note_lines), encoding="utf-8")

    print("Raw ABS folder audit complete.")
    print(f"Target:  {target_dir}")
    print(f"Files:   {len(files)}")
    print(f"Size MB: {run_rows[0]['total_size_mb']}")
    print(f"Duplicates groups: {dup_group_id}")
    print(f"Cleanup candidates: {len(cleanup_rows)}")
    print(f"Inventory: {file_inventory_path}")
    print(f"Cleanup:   {cleanup_path}")
    print(f"Summary:   {run_path}")
    print(f"Note:      {note_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
