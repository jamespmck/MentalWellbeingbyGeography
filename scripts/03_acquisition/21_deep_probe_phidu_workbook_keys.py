#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
21_deep_probe_phidu_workbook_keys.py

Purpose
-------
Deep-probe downloaded PHIDU Social Health Atlas workbooks for usable geography keys.

This script does NOT join PHIDU data into the SA2 master. It scans downloaded PHIDU
Excel workbooks at cell/value level because PHIDU sheets often have multi-row,
merged or duplicated headers that defeat ordinary pandas header detection.

It answers one narrow question:
  Is there a safe SA2, PHN or LGA key in any downloaded PHIDU workbook/sheet that
  can be matched against the v08 MentalWellbeingByGeography master?

Expected input
--------------
  data/processed/integrated/sa2_predictor_universe_v08_with_clean_housing_context.parquet
  data/raw/phidu/*.xlsx

Outputs
-------
  outputs/audits/phidu_deep_key_probe_v11.csv
  outputs/audits/phidu_deep_sheet_summary_v11.csv
  outputs/audits/phidu_deep_join_readiness_v11.csv
  outputs/audits/phidu_deep_probe_run_audit_v11.csv
  docs/source_registers/phidu_deep_key_probe_register_v11.csv
  docs/methodology/phidu_deep_key_probe_note_v11.md

Run
---
  python scripts/03_acquisition/21_deep_probe_phidu_workbook_keys.py --debug
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Any

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise RuntimeError("openpyxl is required to inspect PHIDU Excel workbooks") from exc


DEFAULT_PROJECT_ROOT = Path(r"D:\Good Measure\MentalWellbeingByGeography")
ALT_PROJECT_ROOT = Path(r"D:\Good Measure\MentalWellbeingbyGeography")
DEFAULT_MASTER = Path(r"data\processed\integrated\sa2_predictor_universe_v08_with_clean_housing_context.parquet")
DEFAULT_PHIDU_DIR = Path(r"data\raw\phidu")

KEY_PROBE_COLUMNS = [
    "workbook_path", "workbook_name", "sheet_name", "sheet_index", "column_index", "excel_column",
    "column_label_guess", "target_geography", "target_master_key", "match_mode",
    "unique_source_keys", "matched_target_keys", "match_rate", "matched_sample", "unmatched_sample",
    "non_empty_cells_scanned", "rows_scanned", "recommended_action", "notes",
]

SHEET_SUMMARY_COLUMNS = [
    "workbook_path", "workbook_name", "sheet_name", "sheet_index", "max_row", "max_column",
    "rows_scanned", "columns_scanned", "best_target_geography", "best_target_master_key",
    "best_match_rate", "best_source_column_index", "best_source_excel_column", "best_column_label_guess",
    "recommended_action", "notes",
]

READINESS_COLUMNS = [
    "workbook_path", "workbook_name", "sheet_name", "best_target_geography", "best_target_master_key",
    "best_source_excel_column", "best_column_label_guess", "best_match_rate", "recommended_action", "join_caveat",
]

RUN_AUDIT_COLUMNS = ["check_name", "value", "status", "notes"]


@dataclass
class Logger:
    path: Path
    debug_enabled: bool = False

    def _write(self, level: str, msg: str) -> None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] [{level}] {msg}"
        print(line)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.path.open("a", encoding="utf-8") as f:
            f.write(line + "\n")

    def info(self, msg: str) -> None:
        self._write("INFO", msg)

    def warning(self, msg: str) -> None:
        self._write("WARNING", msg)

    def debug(self, msg: str) -> None:
        if self.debug_enabled:
            self._write("DEBUG", msg)


def resolve_project_root(value: str) -> Path:
    root = Path(value)
    if root.exists():
        return root
    if ALT_PROJECT_ROOT.exists():
        return ALT_PROJECT_ROOT
    return root


def normalise_code(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    s = re.sub(r"[^0-9A-Za-z]", "", s)
    return s.upper()


def normalise_text_key(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).lower().strip()
    s = s.replace("&", " and ")
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\b(primary health network|phn|local government area|lga|city of|shire of|municipality of|regional council)\b", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalise_name(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def excel_col_name(index_1_based: int) -> str:
    out = ""
    n = index_1_based
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def read_master(path: Path, logger: Logger) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Base master not found: {path}")
    logger.info(f"Reading base master: {path}")
    if path.suffix.lower() == ".parquet":
        df = pd.read_parquet(path)
    else:
        df = pd.read_csv(path, dtype=str, low_memory=False)
    logger.info(f"Base master rows: {len(df):,}; columns: {len(df.columns):,}")
    return df


def build_master_targets(master: pd.DataFrame) -> dict[str, dict[str, set[str]]]:
    targets: dict[str, dict[str, set[str]]] = {}
    if "sa2_code_2021" in master.columns:
        targets["SA2"] = {"sa2_code_2021": set(master["sa2_code_2021"].map(normalise_code)) - {""}}
    if "phn_2017_code" in master.columns or "phn_2017_name" in master.columns:
        targets["PHN"] = {}
        if "phn_2017_code" in master.columns:
            targets["PHN"]["phn_2017_code"] = set(master["phn_2017_code"].map(normalise_code)) - {""}
        if "phn_2017_name" in master.columns:
            targets["PHN"]["phn_2017_name"] = set(master["phn_2017_name"].map(normalise_text_key)) - {""}
    if "dominant_lga_code_2021" in master.columns or "dominant_lga_name_2021" in master.columns:
        targets["LGA"] = {}
        if "dominant_lga_code_2021" in master.columns:
            targets["LGA"]["dominant_lga_code_2021"] = set(master["dominant_lga_code_2021"].map(normalise_code)) - {""}
        if "dominant_lga_name_2021" in master.columns:
            targets["LGA"]["dominant_lga_name_2021"] = set(master["dominant_lga_name_2021"].map(normalise_text_key)) - {""}
    return targets


def extract_candidate_tokens(value: Any, target_geography: str, target_key: str) -> list[str]:
    if value is None:
        return []
    try:
        if pd.isna(value):
            return []
    except Exception:
        pass
    s = str(value).strip()
    if not s:
        return []

    if target_key.endswith("_name"):
        key = normalise_text_key(s)
        return [key] if key else []

    # Code matching. Preserve exact normalised cell, plus embedded code tokens.
    out: list[str] = []
    exact = normalise_code(s)
    if exact:
        out.append(exact)

    if target_geography == "SA2":
        out.extend(re.findall(r"(?<!\d)(\d{9})(?!\d)", s))
    elif target_geography == "LGA":
        out.extend(re.findall(r"(?<!\d)(\d{4,5})(?!\d)", s))
    elif target_geography == "PHN":
        # PHN codes can appear as short numeric or PHN-prefixed strings depending on source.
        out.extend(re.findall(r"(?i)\bPHN\s*\d{1,3}\b|(?<!\d)\d{1,3}(?!\d)", s))
        out = [normalise_code(x) for x in out]

    # Deduplicate preserving order.
    seen = set()
    cleaned = []
    for item in out:
        item = normalise_code(item)
        if item and item not in seen:
            seen.add(item)
            cleaned.append(item)
    return cleaned


def column_label_guess(values_by_row: list[Any], max_label_rows: int = 12) -> str:
    labels = []
    for value in values_by_row[:max_label_rows]:
        if value is None:
            continue
        s = str(value).strip()
        if not s or s.lower().startswith("unnamed"):
            continue
        if len(s) > 120:
            s = s[:117] + "..."
        labels.append(s)
    # Use unique labels only.
    seen = set()
    unique = []
    for label in labels:
        key = normalise_name(label)
        if key and key not in seen:
            seen.add(key)
            unique.append(label)
    return " | ".join(unique[:5])


def action_for(target_geography: str, target_master_key: str, match_rate: float, min_match_rate: float) -> tuple[str, str]:
    if match_rate >= min_match_rate:
        if target_geography == "SA2":
            return "safe_candidate_direct_sa2_join_after_indicator_selection", "Validate ASGS year and indicator definitions before joining."
        if target_geography == "PHN":
            return "candidate_phn_context_join_after_boundary_year_confirmation", "PHN context only. Confirm PHIDU PHN boundary year aligns with master PHN context."
        if target_geography == "LGA":
            return "candidate_lga_context_join_after_code_year_validation", "LGA context only. Join through dominant LGA with area-share caveat."
    if match_rate > 0:
        return "manual_review_key_match_below_threshold", "Some key overlap detected but below safe threshold. Inspect unmatched keys."
    return "hold_context_only_no_validated_key_match", "No safe key match detected. Do not join."


def scan_sheet(
    workbook_path: Path,
    workbook_name: str,
    sheet_name: str,
    sheet_index: int,
    rows: list[tuple[Any, ...]],
    max_columns: int,
    targets: dict[str, dict[str, set[str]]],
    min_match_rate: float,
) -> tuple[list[dict], dict]:
    probe_rows: list[dict] = []
    n_rows = len(rows)
    n_cols = min(max((len(r) for r in rows), default=0), max_columns)

    best: dict[str, Any] = {
        "best_match_rate": 0.0,
        "best_target_geography": "",
        "best_target_master_key": "",
        "best_source_column_index": None,
        "best_source_excel_column": "",
        "best_column_label_guess": "",
        "recommended_action": "hold_context_only_no_validated_key_match",
        "notes": "No safe key match detected. Do not join.",
    }

    for col_idx in range(n_cols):
        col_values = [r[col_idx] if col_idx < len(r) else None for r in rows]
        non_empty = sum(1 for v in col_values if v not in (None, ""))
        if non_empty == 0:
            continue
        label = column_label_guess(col_values)

        for target_geography, key_map in targets.items():
            for target_master_key, target_set in key_map.items():
                if not target_set:
                    continue
                source_tokens: list[str] = []
                for value in col_values:
                    source_tokens.extend(extract_candidate_tokens(value, target_geography, target_master_key))
                unique_tokens = sorted(set(source_tokens) - {""})
                if not unique_tokens:
                    continue
                matched = [t for t in unique_tokens if t in target_set]
                match_rate = len(matched) / len(unique_tokens) if unique_tokens else 0.0

                # Avoid noise from columns with just 1 coincidental value unless exact high match.
                if len(unique_tokens) < 2 and match_rate < min_match_rate:
                    continue
                if match_rate == 0 and len(unique_tokens) > 80:
                    # Write fewer zero-match rows; broad numeric indicator columns create noise.
                    continue

                action, notes = action_for(target_geography, target_master_key, match_rate, min_match_rate)
                row = {
                    "workbook_path": str(workbook_path),
                    "workbook_name": workbook_name,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "column_index": col_idx + 1,
                    "excel_column": excel_col_name(col_idx + 1),
                    "column_label_guess": label,
                    "target_geography": target_geography,
                    "target_master_key": target_master_key,
                    "match_mode": "name" if target_master_key.endswith("_name") else "code_or_embedded_code",
                    "unique_source_keys": len(unique_tokens),
                    "matched_target_keys": len(matched),
                    "match_rate": match_rate,
                    "matched_sample": " | ".join(matched[:20]),
                    "unmatched_sample": " | ".join([t for t in unique_tokens if t not in target_set][:30]),
                    "non_empty_cells_scanned": non_empty,
                    "rows_scanned": n_rows,
                    "recommended_action": action,
                    "notes": notes,
                }
                probe_rows.append(row)

                is_better = (
                    match_rate > float(best["best_match_rate"]) or
                    (match_rate == float(best["best_match_rate"]) and len(matched) > 0 and target_geography == "SA2")
                )
                if is_better:
                    best.update({
                        "best_match_rate": match_rate,
                        "best_target_geography": target_geography,
                        "best_target_master_key": target_master_key,
                        "best_source_column_index": col_idx + 1,
                        "best_source_excel_column": excel_col_name(col_idx + 1),
                        "best_column_label_guess": label,
                        "recommended_action": action,
                        "notes": notes,
                    })

    sheet_summary = {
        "workbook_path": str(workbook_path),
        "workbook_name": workbook_name,
        "sheet_name": sheet_name,
        "sheet_index": sheet_index,
        "max_row": n_rows,
        "max_column": n_cols,
        "rows_scanned": n_rows,
        "columns_scanned": n_cols,
        **best,
    }
    return probe_rows, sheet_summary


def iter_sheet_rows(ws, max_rows: int, max_columns: int) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > max_rows:
            break
        if max_columns and len(row) > max_columns:
            row = row[:max_columns]
        rows.append(row)
    return rows


def write_csv_with_schema(rows: list[dict] | pd.DataFrame, path: Path, columns: list[str], logger: Logger) -> pd.DataFrame:
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(rows, pd.DataFrame):
        df = rows.copy()
    else:
        df = pd.DataFrame(rows)
    for col in columns:
        if col not in df.columns:
            df[col] = pd.Series(dtype="object")
    df = df[columns + [c for c in df.columns if c not in columns]]
    logger.info(f"Writing CSV: {path}")
    df.to_csv(path, index=False, encoding="utf-8-sig", quoting=csv.QUOTE_MINIMAL)
    return df


def write_methodology(path: Path, readiness_df: pd.DataFrame, logger: Logger) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# PHIDU Social Health Atlas deep key probe v11",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "This audit performs a cell-level scan of downloaded PHIDU Excel workbooks. It does not join PHIDU data into the SA2 master.",
        "",
        "## Why this step was needed",
        "",
        "PHIDU workbooks use multi-row and sometimes merged or duplicated headers. A conventional pandas header-based scan can miss geography key columns or create empty validation outputs.",
        "",
        "## Decision rule",
        "",
        "A sheet is only treated as a join candidate if one source column or embedded-code column matches at least 95% of unique source keys to the relevant v08 master key. PHN and LGA candidates remain contextual and require boundary/year review even after a high key match.",
        "",
        "## Recommended actions summary",
        "",
    ]
    if not readiness_df.empty and "recommended_action" in readiness_df.columns:
        summary = readiness_df["recommended_action"].value_counts(dropna=False).reset_index()
        summary.columns = ["recommended_action", "sheet_count"]
        lines.append(summary.to_markdown(index=False))
    else:
        lines.append("No join-ready sheets were found.")
    lines.extend([
        "",
        "## Key caution",
        "",
        "Do not broad-join PHIDU data. Select a small set of indicators, validate geography, validate denominator definitions, then add a source-specific context layer.",
    ])
    logger.info(f"Writing methodology note: {path}")
    path.write_text("\n".join(lines), encoding="utf-8")


def notify(success: bool, title: str, detail: str = "") -> None:
    try:
        if sys.platform.startswith("win"):
            import winsound
            winsound.MessageBeep(winsound.MB_ICONASTERISK if success else winsound.MB_ICONHAND)
            msg = f"{title}\n{detail}"[:900].replace("'", "’")
            icon = 64 if success else 16
            ps = "$wshell = New-Object -ComObject WScript.Shell; " + f"$null = $wshell.Popup('{msg}', 12, '{title}', {icon})"
            subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=17, check=False)
        else:
            print("\a", end="", flush=True)
    except Exception:
        pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Deep-probe downloaded PHIDU workbooks for join keys.")
    parser.add_argument("--project-root", default=str(DEFAULT_PROJECT_ROOT))
    parser.add_argument("--base-master", default=str(DEFAULT_MASTER))
    parser.add_argument("--phidu-dir", default=str(DEFAULT_PHIDU_DIR))
    parser.add_argument("--max-rows-per-sheet", type=int, default=3500)
    parser.add_argument("--max-columns-per-sheet", type=int, default=220)
    parser.add_argument("--max-workbooks", type=int, default=0, help="0 means all workbooks")
    parser.add_argument("--max-sheets-per-workbook", type=int, default=0, help="0 means all sheets")
    parser.add_argument("--min-match-rate", type=float, default=0.95)
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = resolve_project_root(args.project_root)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    logger = Logger(root / "outputs" / "logs" / f"21_deep_probe_phidu_workbook_keys_{ts}.log", debug_enabled=args.debug)

    logger.info("PHIDU deep workbook key probe v11")
    logger.info(f"Project root: {root}")
    logger.info(f"Log path: {logger.path}")

    master_path = Path(args.base_master)
    if not master_path.is_absolute():
        master_path = root / master_path
    phidu_dir = Path(args.phidu_dir)
    if not phidu_dir.is_absolute():
        phidu_dir = root / phidu_dir

    master = read_master(master_path, logger)
    targets = build_master_targets(master)
    target_summary = ", ".join(f"{geo}:{';'.join(k + '=' + str(len(v)) for k, v in keys.items())}" for geo, keys in targets.items())
    logger.info("Target key sets: " + target_summary)

    if not phidu_dir.exists():
        raise FileNotFoundError(f"PHIDU download directory not found: {phidu_dir}")
    workbooks = sorted([p for p in phidu_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    if args.max_workbooks and args.max_workbooks > 0:
        workbooks = workbooks[:args.max_workbooks]
    logger.info(f"PHIDU workbooks selected for deep probe: {len(workbooks):,}")

    probe_rows: list[dict] = []
    summary_rows: list[dict] = []
    failed_workbooks: list[str] = []
    sheets_scanned = 0

    for wb_i, workbook_path in enumerate(workbooks, start=1):
        logger.info(f"Opening workbook {wb_i}/{len(workbooks)}: {workbook_path.name}")
        try:
            wb = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"Could not open workbook {workbook_path}: {type(exc).__name__}: {exc}")
            failed_workbooks.append(str(workbook_path))
            continue

        sheet_names = list(wb.sheetnames)
        if args.max_sheets_per_workbook and args.max_sheets_per_workbook > 0:
            sheet_names = sheet_names[:args.max_sheets_per_workbook]
        logger.info(f"  Sheets selected: {len(sheet_names):,}")

        for sheet_i, sheet_name in enumerate(sheet_names, start=1):
            try:
                ws = wb[sheet_name]
                rows = iter_sheet_rows(ws, args.max_rows_per_sheet, args.max_columns_per_sheet)
                p_rows, s_row = scan_sheet(
                    workbook_path=workbook_path,
                    workbook_name=workbook_path.name,
                    sheet_name=sheet_name,
                    sheet_index=sheet_i,
                    rows=rows,
                    max_columns=args.max_columns_per_sheet,
                    targets=targets,
                    min_match_rate=args.min_match_rate,
                )
                probe_rows.extend(p_rows)
                summary_rows.append(s_row)
                sheets_scanned += 1
                if args.debug and s_row.get("best_match_rate", 0) and float(s_row.get("best_match_rate", 0)) > 0:
                    logger.debug(f"  {sheet_name}: best {s_row['best_target_geography']} {s_row['best_target_master_key']} rate={s_row['best_match_rate']:.3f} col={s_row['best_source_excel_column']}")
            except Exception as exc:  # noqa: BLE001
                logger.warning(f"Could not scan sheet {workbook_path.name}::{sheet_name}: {type(exc).__name__}: {exc}")
                summary_rows.append({
                    "workbook_path": str(workbook_path),
                    "workbook_name": workbook_path.name,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_i,
                    "max_row": "",
                    "max_column": "",
                    "rows_scanned": 0,
                    "columns_scanned": 0,
                    "best_target_geography": "",
                    "best_target_master_key": "",
                    "best_match_rate": 0.0,
                    "best_source_column_index": "",
                    "best_source_excel_column": "",
                    "best_column_label_guess": "",
                    "recommended_action": "scan_failed_hold_context_only",
                    "notes": f"{type(exc).__name__}: {exc}",
                })
                continue

        try:
            wb.close()
        except Exception:
            pass

    out_audit = root / "outputs" / "audits"
    out_docs = root / "docs" / "source_registers"
    out_methods = root / "docs" / "methodology"

    probe_df = write_csv_with_schema(probe_rows, out_audit / "phidu_deep_key_probe_v11.csv", KEY_PROBE_COLUMNS, logger)
    summary_df = write_csv_with_schema(summary_rows, out_audit / "phidu_deep_sheet_summary_v11.csv", SHEET_SUMMARY_COLUMNS, logger)

    readiness_df = summary_df.copy()
    if not readiness_df.empty:
        readiness_df = readiness_df[readiness_df["recommended_action"].astype(str).str.contains("candidate|review", case=False, na=False)].copy()
        readiness_df = readiness_df.sort_values(["best_match_rate", "best_target_geography"], ascending=[False, True])
        readiness_df = readiness_df.rename(columns={"notes": "join_caveat"})
    readiness_df = write_csv_with_schema(readiness_df, out_audit / "phidu_deep_join_readiness_v11.csv", READINESS_COLUMNS, logger)
    write_csv_with_schema(readiness_df, out_docs / "phidu_deep_key_probe_register_v11.csv", READINESS_COLUMNS, logger)

    run_rows = [
        {"check_name": "base_master_file", "value": str(master_path), "status": "info", "notes": "v08 master used for key matching."},
        {"check_name": "base_master_rows", "value": len(master), "status": "pass" if len(master) == 2472 else "review", "notes": "Expected SA2 row count is 2472."},
        {"check_name": "phidu_download_directory", "value": str(phidu_dir), "status": "pass" if phidu_dir.exists() else "fail", "notes": "Downloaded PHIDU workbook directory."},
        {"check_name": "workbooks_scanned", "value": len(workbooks), "status": "pass" if len(workbooks) > 0 else "review", "notes": "Excel workbooks selected for deep scan."},
        {"check_name": "failed_workbooks", "value": len(failed_workbooks), "status": "pass" if not failed_workbooks else "review", "notes": " | ".join(failed_workbooks[:10])},
        {"check_name": "sheets_scanned", "value": sheets_scanned, "status": "pass" if sheets_scanned > 0 else "review", "notes": "Workbook sheets scanned at cell level."},
        {"check_name": "probe_rows", "value": len(probe_df), "status": "info", "notes": "Column/key probe rows written."},
        {"check_name": "readiness_rows", "value": len(readiness_df), "status": "info", "notes": "Sheets with candidate/review key evidence."},
    ]
    write_csv_with_schema(run_rows, out_audit / "phidu_deep_probe_run_audit_v11.csv", RUN_AUDIT_COLUMNS, logger)
    write_methodology(out_methods / "phidu_deep_key_probe_note_v11.md", readiness_df, logger)

    logger.info("PHIDU deep key probe complete.")
    if not readiness_df.empty:
        logger.info("Recommended actions:\n" + readiness_df["recommended_action"].value_counts(dropna=False).to_string())
    else:
        logger.info("No join-ready or review candidate sheets found.")
    logger.info("Next action: inspect phidu_deep_join_readiness_v11.csv and phidu_deep_key_probe_v11.csv before any PHIDU integration script.")
    notify(True, "PHIDU deep key probe completed", f"Sheets scanned: {sheets_scanned}; readiness rows: {len(readiness_df)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # noqa: BLE001
        notify(False, "PHIDU deep key probe failed", f"{type(exc).__name__}: {exc}")
        raise
